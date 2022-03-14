import random as rvar
import datetime as dt_var
import openpyxl as xl

game_rounds=0

date_and_time=dt_var.datetime.now()
#board is a list type variable that displays the game board for playing tic tac toe
board=["_","_","_",
       "_","_","_",
       "_","_","_"]
#A player can select either X or O to place on the game board
game_var=["X", "O"]

#These are the positions on the game board where a player can insert either X or O

board_position_var=[1,2,3,4,5,6,7,8,9]

# the game_board function will display the game board on the screen
def game_board():
    print()
    print(board[0],"|",board[1],"|",board[2])
    print(board[3],"|",board[4],"|",board[5])
    print(board[6],"|",board[7],"|",board[8])
    print()

# game toss function conducts a toss to decide which player will make the first move
def game_toss(player1,player2):
    var1=rvar.randint(1,10)
    var2=rvar.randint(1,10)
    tmp=" "
    if(var1==var2):
        game_toss(player1,player2)
    elif(var1>var2):
        print(player1," Won the toss")
        print()
        return 1
    else:
        print(player2," Won the toss")
        print()
        return 2


# this function helps the toss winner to select his favourite game vairable that is either X or O
def game_var_select(toss_var,player):
    print("congratulations ",player ,"on winning the toss ","first move is yours to make")
    print(player,"Select your gaming variable")
    print("Press 1 to select X")
    print("Press 2 to select O")
    var=int(input())
    tmp=" "
    if toss_var==1:
        if (var == 2):
            tmp = game_var[0]
            game_var[0] = game_var[1]
            game_var[1] = tmp
            print("Game variable ", game_var[0], " succesfully selected")
        elif var == 1:
            print("Game variable ", game_var[0], " succesfully selected")
        else:
            print("You pressed the wrong selection please select again")
            game_var_select(toss_var, player)
    else:
        if (var == 1):
            tmp = game_var[1]
            game_var[1] = game_var[0]
            game_var[0] = tmp
            print("Game variable ", game_var[1], " succesfully selected")
        elif var == 2:
            print("Game variable ", game_var[1], " succesfully selected")
        else:
            print("You pressed the wrong selection please select again")
            game_var_select(toss_var, player)


# this function allows the current player to make his/her move on the game board
def player_turn(current_player):
    print(current_player[1],"Make your move")
    print("The folowing are availible positions to insert you game variable ",board_position_var)
    move_var=int(input("Enter postion number to make your move -"))
    try:
        board_position_var.remove(move_var)
        board[move_var - 1] = game_var[current_player[0] - 1]
        game_board()
    except :
        print()
        print("You might have entered incorrect position number")
        print(current_player[1]," Please select again")
        print()
        player_turn(current_player)


#game logic of tic tac toe begins here
def game_check():
    if diagnol() or row() or column():
        return 0
    else:
        return 1

def row():
    row1 = board[0] == board[1] == board[2] !="_"
    row2 = board[3] == board[4] == board[5] !="_"
    row3 = board[6] == board[7] == board[8] !="_"
    if row1 or row2 or row3:
        return 1
    else:
        return 0

def column():
    column1 = board[0] == board[3] == board[6] != "_"
    column2 = board[1] == board[4] == board[7] != "_"
    column3 = board[2] == board[5] == board[8] != "_"
    if column1 or column2 or column3:
        return 1
    else:
        return 0
def diagnol():
    diagnol1 = board[0] == board[4] == board[8] != "_"
    diagnol2 = board[2] == board[4] == board[6] != "_"
    if diagnol1 or diagnol2:
        return 1
    else:
        return 0
# game logic of tic tac toe ends here

#this function helps to change player turns
def player_change(current,player1,player2):
    if current[1]==player1[1]:
        current=player2.copy()
    else:
        current=player1.copy()
    return current


#this function will reset the game board and game positon variables
def game_reset():
    global board_position_var, board
    board_position_var = [1, 2, 3, 4, 5, 6, 7, 8, 9].copy()
    board = ["_", "_", "_",
             "_", "_", "_",
             "_", "_", "_"].copy()


#this function declares the result of the game
def game_result(current_player,game_check_var,tie_check,data):
    if game_check_var==0 and tie_check==0:
        game_reset()
        game_log(game_rounds,data,current_player,tie_check)
        print("congratulations ",current_player[1]," you have won")
        print("Press 1 to play again")
        print("Press 2 to exit")
        check=int(input())
        if(check==1):
            start_game()
        else:
            print("thank you for playing tic tac toe")

    elif game_check_var==1 and tie_check==1:
        game_reset()
        print("Well played both players its a tie")
        print("congratulations ", current_player[1], " you have won")
        print("Press 1 to play again")
        print("Press 2 to exit")
        check = int(input())
        if (check == 1):
            start_game()
        else:
            print("thank you for playing tic tac toe")


#this function generates the game log
def game_log(game_rounds,data,winner,tie_check):
    game_log_file=open("Game_logs.txt","a")
    game_rounds=game_rounds+1
    tie_msg=False
    if tie_check==1:
        tie_msg=True
    game_log_file.write((str(game_rounds)+" "+data[1]+" "+data[3]+" "+winner[1]+" "+str(tie_msg))+" "+str(date_and_time)+"\n")
    game_log_file.close()
    game_record()


#this function keeps the record of the game
def game_record():
    wb = xl.Workbook()
    ws = wb.active
    ws.cell(row=1,column=1).value="Game rounds"
    ws.cell(row=1, column=2).value ="Player 1 name"
    ws.cell(row=1,column=3).value="Player 2 name"
    ws.cell(row=1,column=4).value="Winner"
    ws.cell(row=1,column=5).value="Tie check"
    ws.cell(row=1, column=6).value = "Date"
    ws.cell(row=1, column=7).value = "Time"

    game_log_file=open("Game_logs.txt","r")
    row_var=2
    col_var=1
    row_count=0
    for x in game_log_file:
        tmp=x.split()
        for y in range(0,7):
            ws.cell(row=row_var + row_count, column=col_var + y).value = tmp[y]
        row_count=row_count+1
    wb.save("Game_records.xlsx")




#this is the game controller function that calls every important function for playing the game
def play_game(player1,player2):
    print("\n","-----Starting tic tac toe-----")
    print("---",player1," VS ",player2,"---")
    print("---Time to conduct a toss---")
    print()
    toss_result=game_toss(player1,player2)
    current_player=list()
    player1_data=[1,player1]
    player2_data=[2,player2]
    data=player1_data+player2_data
    game_check_var=1
    tie_check=0
    if toss_result==1:
        current_player=player1_data.copy()
        game_var_select(toss_result,player1)
    else:
        current_player =player2_data.copy()
        game_var_select(toss_result, player2)

    while game_check_var:
        if board_position_var==[]:
            tie_check=1
            break
        player_turn(current_player)
        game_check_var = game_check()
        if game_check_var and board_position_var!=[]:
            current_player = player_change(current_player, player1_data, player2_data)

    game_result(current_player,game_check_var,tie_check,data)





#this is the starter function that starts the game and displays all the important messages
def start_game():

    print("---Tic Tac Toe---")
    game_board()
    print("---Welcome to the game---")
    print("---Menu---","\n","To start game press 1","\n","To exit press 0")
    testVar=int(input())
    if testVar==1:
        player1=input("Enter 1st player name- ")
        tmp1=player1.split()
        player1="_".join(tmp1)
        player2 =input("Enter 2nd  player name- ")
        tmp2 = player2.split()
        player2 = "_".join(tmp2)
        play_game(player1,player2)
    elif testVar==0:
        print("Thank you for playing tic tac toe")
    else:
        print("Wrong option please select Again")
        start_game()


start_game()
