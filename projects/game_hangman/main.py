# Words.py is a dependency file to the game that contains a list of all mystery words in the game
# Words.py contains only a list of words
# Word.py is very essential for this program to execute successfully
from words import word_list
import datetime
import random
import openpyxl as xl
date_time=datetime.datetime.now()

game_status=False
game_rounds=0

def game_control(): #this function controls the flow of game once the game result has been generated
    print()
    print("Press 1 to play again")
    print("Press 2 to exit")
    menu_var=0
    try:
        menu_var = int(input())
    except:
        print("Invalid selection, please select the correct option from the menu")
        game_control()
    if menu_var==1:
        start_game()
    elif menu_var==2:
        print("Thank you for playing hangman")
    else:
        print("Invalid input, please select the right option from the menu")
        game_control()


#this function generates a log txt file of the game that contains information about players and game result
def game_log(player):
    global game_rounds,game_status
    game_rounds=game_rounds+1
    game_file=open("Game_log.txt","a")
    result="LOSS"
    if game_status:
        result="WIN"
    game_file.write(str(game_rounds)+" "+player+" "+result+" "+str(date_time)+"\n")
    game_file.close()
    game_records()
    game_status=False


# this function generates an excel file that keeps the record of all players and their game results
def game_records():
    game_file=open("Game_log.txt","r")
    wb=xl.Workbook()
    ws=wb.active
    ws.cell(row=1,column=1).value  = "Game rounds"
    ws.cell(row=1, column=2).value = "Player name"
    ws.cell(row=1, column=3).value = "Game status"
    ws.cell(row=1, column=4).value = "Date"
    ws.cell(row=1, column=5).value = "Time"
    row_var=2
    col_var=1
    for x in game_file:
        tmp=x.split()
        for y in range(0,5):
            ws.cell(row=row_var,column=col_var+y).value=tmp[y]
        row_var=row_var+1
    game_file.close()
    wb.save("Game_records.xlsx")


# this function returns the current status of the hangman
def current_state(no_of_turns):
    # comments are used as list items in h_man to store  all the states of hanging man
    h_man = [  # this is the final state of hanging man with all body parts: head, torso, both arms, and both legs
        """
           --------
           |      |
           |      O
           |     \\|/
           |      |
           |     / \\
           -
        """,
        # this is the 5 th state  state of hanging man with the following parts: head, torso, both arms, and one leg
        """
           --------
           |      |
           |      O
           |     \\|/
           |      |
           |     / 
           -
        """,
        # this is the 4 th state  state of hanging man with the following parts:head, torso, and both arms
        """
           --------
           |      |
           |      O
           |     \\|/
           |      |
           |      
           -
        """,
        # this is the 3rd state  state of hanging man with the following parts:head, torso, and one arm
        """
           --------
           |      |
           |      O
           |     \\|
           |      |
           |     
           -
        """,
        # this is the 2nd state of hanging man with the following parts:head and torso
        """
           --------
           |      |
           |      O
           |      |
           |      |
           |     
           -
        """,
        # this is the 1 st state of hanging man with the following parts:head
        """
           --------
           |      |
           |      O
           |    
           |      
           |     
           -
        """,
        # this is the initial state of hanging man
        """
           --------
           |      |
           |      
           |    
           |      
           |     
           -
        """
    ]
    return h_man[no_of_turns]


#this function fetches the mystery words from the game dependency file words.py
def fetch_word():
    rand_word=random.choice(word_list)
    word=rand_word.upper()
    return word

#this function is the main function of the game
#play_game function is responsible for computing the game logic succesully and call other support functions
def play_game(word,player):
    word_completion = "_ " * len(word)
    guessed_word=[]
    guessed_letter=[]
    word_guess = False
    no_of_turns = 6


    # Game logic starts here
    while not word_guess and no_of_turns>0:
        print(current_state(no_of_turns))
        print("Mystery Word : ",word_completion)
        print()
        guess=input("Guess a letter or word : ").upper().strip()
        if len(guess)==1 and guess.isalpha():
            if guess in guessed_letter:
                print("The letter has been already guessed")
            elif guess not in word:
                print("The aplphabet ",guess," is not present in the word")
                guessed_letter.append(guess)
                no_of_turns=no_of_turns-1
            else:
                print("Wow you made a correct guess")
                guessed_letter.append(guess)
                word_completion_list = word_completion.split()
                tmp_word_list = []
                for x in word:
                    tmp_word_list.append(x)
                for y in range(0,len(word)):
                    if (tmp_word_list[y] == guess):
                        word_completion_list[y] = guess
                word_completion = " ".join(word_completion_list)
                if "_" not in word_completion:
                    word_guess=True

        elif len(guess)==len(word) and guess.isalpha():
            if guess in guessed_word:
                print("Woahhh the word was already guessed")
            elif guess!=word:
                print("How unfortunate, your guess word is wrong")
                guessed_word.append(guess)
                no_of_turns=no_of_turns-1
            else:
                print("Great job, You guessed the correct word")
                word_guess=True
                word_completion=word
        else:
            print("Hahahaha incorrect guess")
            print("Well you gotta try harder if you want to save the man")
            no_of_turns=no_of_turns-1
   # Game logic ends here


    if word_guess==True:
        global game_status
        game_status=True
        game_log(player)
        print("Congratulations",player," you win,")
        game_control()

    else:
        game_log(player)
        print("Better luck next time, the mystery word was ",word)
        game_control()


# this function is the starter function of the game it prints the necessary output functions on the console
def start_game():

    print("-----------------HANG MAN------------------")
    print(current_state(0))
    print("Press 1 to play")
    print("Press 2 to exit")
    menu_choice=""
    try:
        menu_choice = int(input())
    except:
        print("Invaild Input please select appropriate option")
        start_game()
    if(menu_choice==1):
        print("Enter player name - ")
        player_name=input().strip()
        print(player_name," You have 6 turns to guess the correct word, select wisely - ")
        tmp=player_name.split()
        player_name="_".join(tmp)

        play_game(fetch_word(),player_name)
    elif(menu_choice==2):
        print("Thank you for playing hangman")
    else:
        print("Invalid input please select the appropraite option")
        print()
        start_game()


start_game()

